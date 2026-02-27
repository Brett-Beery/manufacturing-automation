import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from data import df_production
from log_gen import create_log_template
import datetime
import os

def get_user_inputs():
    print("\n=== PROCESS LOG GENERATOR ===\n")
    
    # Product selection
    products = df_production["Product"].unique().tolist()
    print("Available Products:")
    for i, product in enumerate(products, 1):
        print(f"  {i}. {product}")
    
    while True:
        try:
            product_choice = int(input("\nSelect product number: ")) - 1
            if 0 <= product_choice < len(products):
                selected_product = products[product_choice]
                break
            else:
                print("Invalid selection, please try again.")
        except ValueError:
            print("Please enter a number.")

    # Line speed selection
    speeds = df_production[df_production["Product"] == selected_product]["Line_Speed"].tolist()
    print(f"\nAvailable Line Speeds for {selected_product}:")
    for i, speed in enumerate(speeds, 1):
        print(f"  {i}. {speed}")

    while True:
        try:
            speed_choice = int(input("\nSelect line speed number: ")) - 1
            if 0 <= speed_choice < len(speeds):
                selected_speed = speeds[speed_choice]
                break
            else:
                print("Invalid selection, please try again.")
        except ValueError:
            print("Please enter a number.")

    # Additional info
    operator = input("\nOperator name: ")
    shift = input("Shift (1/2/3): ")
    work_order = input("Work Order #: ")
    date = datetime.date.today().strftime("%m/%d/%Y")

    return selected_product, selected_speed, operator, shift, work_order, date


def populate_log(product, speed, operator, shift, work_order, date):
    # Pull specs from dataframe
    specs = df_production[
        (df_production["Product"] == product) &
        (df_production["Line_Speed"] == speed)
    ].iloc[0]

    # Generate fresh template
    template_path = "process_log_template.xlsx"
    create_log_template(template_path)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Process Log"]

    # --- Populate Info Block ---
    ws["B3"] = product
    ws["B4"] = speed
    ws["B5"] = date
    ws["B6"] = shift
    ws["B7"] = operator
    ws["B8"] = work_order

    # --- Raw Material Checks ---
    rm_checks = [
        ("Primary Material", specs["Primary_Material"], "-", "-"),
        ("Material Spec", specs["Primary_Material_Spec"], "-", "-"),
        ("Adhesive Type", specs["Adhesive"], "-", "-"),
        ("Adhesive Rate (gpm)", specs["Adhesive_Rate_gpm"], 
         specs["Adhesive_Rate_gpm"] * 0.95, specs["Adhesive_Rate_gpm"] * 1.05),
        ("Thickness (mil)", "-", specs["Thickness_Min_mil"], specs["Thickness_Max_mil"]),
    ]

    rm_start_row = 12
    for i, (check, spec, min_val, max_val) in enumerate(rm_checks):
        row = rm_start_row + i
        ws.cell(row=row, column=1, value=check)
        ws.cell(row=row, column=2, value=spec)
        ws.cell(row=row, column=3, value=min_val)
        ws.cell(row=row, column=4, value=max_val)

    # --- Machine Parameter Checks ---
    mp_checks = [
        ("Temperature (F)", specs["Temperature_F"], 
         specs["Temperature_F"] - 10, specs["Temperature_F"] + 10),
        ("Pressure (psi)", specs["Pressure_psi"],
         specs["Pressure_psi"] - 3, specs["Pressure_psi"] + 3),
        ("Line Speed (fpm)", specs["Line_Speed_fpm"],
         specs["Line_Speed_fpm"] - 10, specs["Line_Speed_fpm"] + 10),
        ("Setup Time (min)", specs["Setup_Time_min"], "-", "-"),
        ("Post Processing", specs["Post_Processing"], "-", "-"),
    ]

    mp_start_row = 20
    for i, (param, spec, min_val, max_val) in enumerate(mp_checks):
        row = mp_start_row + i
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=spec)
        ws.cell(row=row, column=3, value=min_val)
        ws.cell(row=row, column=4, value=max_val)

    # --- Save ---
    output_filename = f"ProcessLog_{product.replace(' ', '_')}_{speed}_{date.replace('/', '-')}.xlsx"
    wb.save(output_filename)
    print(f"\nProcess log saved as: {output_filename}")
    return output_filename


if __name__ == "__main__":
    product, speed, operator, shift, work_order, date = get_user_inputs()
    output = populate_log(product, speed, operator, shift, work_order, date)
    print("\nDone! Open your process log to review.")