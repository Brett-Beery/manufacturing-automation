import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from data import df_production, df_labels
import datetime
import math

# --- Styles ---
def make_styles():
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F8F")
    section_font = Font(name="Calibri", bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    normal_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header_font": header_font,
        "header_fill": header_fill,
        "section_font": section_font,
        "section_fill": section_fill,
        "normal_font": normal_font,
        "bold_font": bold_font,
        "center": center,
        "left": left,
        "right": right,
        "border": border
    }

def get_user_inputs():
    print("\n=== WORK ORDER GENERATOR ===\n")

    # Product selection
    products = df_production["Product"].unique().tolist()
    print("Available Products:")
    for i, product in enumerate(products, 1):
        print(f"  {i}. {product}")

    while True:
        try:
            choice = int(input("\nSelect product number: ")) - 1
            if 0 <= choice < len(products):
                selected_product = products[choice]
                break
            else:
                print("Invalid selection, please try again.")
        except ValueError:
            print("Please enter a number.")

    # Line speed selection
    speeds = df_production[
        df_production["Product"] == selected_product
    ]["Line_Speed"].tolist()
    print(f"\nAvailable Line Speeds for {selected_product}:")
    for i, speed in enumerate(speeds, 1):
        print(f"  {i}. {speed}")

    while True:
        try:
            choice = int(input("\nSelect line speed number: ")) - 1
            if 0 <= choice < len(speeds):
                selected_speed = speeds[choice]
                break
            else:
                print("Invalid selection, please try again.")
        except ValueError:
            print("Please enter a number.")

    # Work order details
    work_order = input("\nWork Order #: ")
    start_date = input("Start Date (MM/DD/YYYY): ")
    end_date = input("End Date (MM/DD/YYYY): ")
    order_qty = int(input("Order Quantity: "))
    shift_target = int(input("Shift Target Quantity: "))

    return selected_product, selected_speed, work_order, start_date, end_date, order_qty, shift_target


def build_work_order(product, speed, work_order, start_date, end_date, order_qty, shift_target):
    specs = df_production[
        (df_production["Product"] == product) &
        (df_production["Line_Speed"] == speed)
    ].iloc[0]

    label_specs = df_labels[df_labels["Product"] == product].iloc[0]

    s = make_styles()
    wb = openpyxl.Workbook()

    # =====================
    # SHEET 1 - WORK ORDER
    # =====================
    ws = wb.active
    ws.title = "Work Order"

    # Column widths
    col_widths = [22, 20, 15, 15, 15, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "WORK ORDER"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = s["header_fill"]
    ws["A1"].alignment = s["center"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Manufacturing Automation Demo"
    ws["A2"].font = Font(name="Calibri", bold=True, size=12)
    ws["A2"].alignment = s["center"]

    # Title block
    title_data = [
        ("Work Order #:", work_order),
        ("Product:", product),
        ("Line Speed:", speed),
        ("Start Date:", start_date),
        ("End Date:", end_date),
        ("Order Quantity:", f"{order_qty:,}"),
        ("Shift Target:", f"{shift_target:,}"),
    ]

    for i, (label, value) in enumerate(title_data):
        row = i + 3
        ws.cell(row=row, column=1, value=label).font = s["bold_font"]
        ws.cell(row=row, column=1).alignment = s["left"]
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = s["normal_font"]
        cell.border = s["border"]
        ws.row_dimensions[row].height = 18

    # Production calculations
    units_per_shift = specs["Units_Per_Shift"]
    units_per_hour = specs["Units_Per_Hour"]
    units_per_day = units_per_hour * 8 * 3  # 3 shifts per day
    total_shifts = math.ceil(order_qty / shift_target)
    total_days = math.ceil(total_shifts / 3)

    calc_start = 11
    ws.merge_cells(f"A{calc_start}:F{calc_start}")
    ws[f"A{calc_start}"] = "PRODUCTION SUMMARY"
    ws[f"A{calc_start}"].font = s["section_font"]
    ws[f"A{calc_start}"].fill = s["section_fill"]
    ws[f"A{calc_start}"].alignment = s["center"]

    calc_data = [
        ("Units Per Hour:", f"{units_per_hour:,}"),
        ("Units Per Shift:", f"{units_per_shift:,}"),
        ("Units Per Day:", f"{units_per_day:,}"),
        ("Total Shifts Required:", f"{total_shifts}"),
        ("Estimated Days to Complete:", f"{total_days}"),
    ]

    for i, (label, value) in enumerate(calc_data):
        row = calc_start + i + 1
        ws.cell(row=row, column=1, value=label).font = s["bold_font"]
        ws.cell(row=row, column=2, value=value).border = s["border"]
        ws.row_dimensions[row].height = 18

    # BOM Section
    bom_start = calc_start + len(calc_data) + 3
    ws.merge_cells(f"A{bom_start}:F{bom_start}")
    ws[f"A{bom_start}"] = "BILL OF MATERIALS"
    ws[f"A{bom_start}"].font = s["section_font"]
    ws[f"A{bom_start}"].fill = s["section_fill"]
    ws[f"A{bom_start}"].alignment = s["center"]

    bom_headers = ["Material", "Specification", "Type", "Rate", "Unit", "Notes"]
    for col, header in enumerate(bom_headers, 1):
        cell = ws.cell(row=bom_start + 1, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["border"]

    bom_items = [
        (specs["Primary_Material"], specs["Primary_Material_Spec"], "Primary", "-", "Roll", "-"),
        (specs["Adhesive"], "-", "Adhesive", specs["Adhesive_Rate_gpm"], "gpm", "-"),
    ]

    for i, item in enumerate(bom_items):
        row = bom_start + 2 + i
        for col, val in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = s["normal_font"]
            cell.border = s["border"]
            ws.row_dimensions[row].height = 18

    # Post Processing
    pp_start = bom_start + len(bom_items) + 4
    ws.merge_cells(f"A{pp_start}:F{pp_start}")
    ws[f"A{pp_start}"] = "POST PROCESSING INSTRUCTIONS"
    ws[f"A{pp_start}"].font = s["section_font"]
    ws[f"A{pp_start}"].fill = s["section_fill"]
    ws[f"A{pp_start}"].alignment = s["center"]

    ws.merge_cells(f"A{pp_start+1}:B{pp_start+1}")
    ws[f"A{pp_start+1}"] = "Process:"
    ws[f"A{pp_start+1}"].font = s["bold_font"]
    ws.merge_cells(f"C{pp_start+1}:F{pp_start+1}")
    ws[f"C{pp_start+1}"] = specs["Post_Processing"]
    ws[f"C{pp_start+1}"].border = s["border"]

    ws.merge_cells(f"A{pp_start+2}:B{pp_start+2}")
    ws[f"A{pp_start+2}"] = "Notes:"
    ws[f"A{pp_start+2}"].font = s["bold_font"]
    ws.merge_cells(f"C{pp_start+2}:F{pp_start+2}")
    ws[f"C{pp_start+2}"] = specs["Post_Process_Notes"]
    ws[f"C{pp_start+2}"].border = s["border"]

    # =====================
    # SHEET 2 - SETUP SHEET
    # =====================
    ws2 = wb.create_sheet("Setup Sheet")

    col_widths2 = [25, 20, 15, 15, 20]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "MACHINE SETUP SHEET"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws2["A1"].fill = s["header_fill"]
    ws2["A1"].alignment = s["center"]
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:E2")
    ws2["A2"] = f"{product} | {speed} Speed | WO# {work_order}"
    ws2["A2"].font = Font(name="Calibri", bold=True, size=12)
    ws2["A2"].alignment = s["center"]

    setup_headers = ["Parameter", "Set Point", "Min", "Max", "Verified"]
    for col, header in enumerate(setup_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["border"]

    setup_params = [
        ("Temperature (F)", specs["Temperature_F"],
         specs["Temperature_F"] - 10, specs["Temperature_F"] + 10),
        ("Pressure (psi)", specs["Pressure_psi"],
         specs["Pressure_psi"] - 3, specs["Pressure_psi"] + 3),
        ("Line Speed (fpm)", specs["Line_Speed_fpm"],
         specs["Line_Speed_fpm"] - 10, specs["Line_Speed_fpm"] + 10),
        ("Adhesive Rate (gpm)", specs["Adhesive_Rate_gpm"],
         round(specs["Adhesive_Rate_gpm"] * 0.95, 2),
         round(specs["Adhesive_Rate_gpm"] * 1.05, 2)),
        ("Setup Time (min)", specs["Setup_Time_min"], "-", "-"),
        ("Changeover Time (min)", specs["Changeover_Time_min"], "-", "-"),
    ]

    for i, (param, setpoint, min_val, max_val) in enumerate(setup_params):
        row = 4 + i
        ws2.cell(row=row, column=1, value=param).font = s["normal_font"]
        ws2.cell(row=row, column=2, value=setpoint).border = s["border"]
        ws2.cell(row=row, column=3, value=min_val).border = s["border"]
        ws2.cell(row=row, column=4, value=max_val).border = s["border"]
        ws2.cell(row=row, column=5, value="").border = s["border"]
        ws2.row_dimensions[row].height = 18

    # =====================
    # SHEET 3 - LABEL INFO
    # =====================
    ws3 = wb.create_sheet("Label Info")

    col_widths3 = [25, 25, 20, 20]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "LABEL INFORMATION"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws3["A1"].fill = s["header_fill"]
    ws3["A1"].alignment = s["center"]
    ws3.row_dimensions[1].height = 30

    ws3.merge_cells("A2:D2")
    ws3["A2"] = f"{product} | Customer: {label_specs['Customer']}"
    ws3["A2"].font = Font(name="Calibri", bold=True, size=12)
    ws3["A2"].alignment = s["center"]

    label_data = [
        ("Customer:", label_specs["Customer"]),
        ("Finished Width (in):", label_specs["Finished_Width_in"]),
        ("Finished Length (in):", label_specs["Finished_Length_in"]),
        ("Roll Weight (lbs):", label_specs["Roll_Weight_lbs"]),
        ("Label 1 - Primary ID:", label_specs["Label_1"]),
        ("Label 2 - Lot Code:", label_specs["Label_2"]),
        ("Label 3 - Customer Spec:", label_specs["Label_3"]),
        ("Label 4 - Shipping:", label_specs["Label_4"]),
        ("Hazmat:", str(label_specs["Hazmat"])),
        ("Special Instructions:", label_specs["Special_Instructions"]),
    ]

    for i, (label, value) in enumerate(label_data):
        row = i + 3
        ws3.cell(row=row, column=1, value=label).font = s["bold_font"]
        cell = ws3.cell(row=row, column=2, value=value)
        cell.font = s["normal_font"]
        cell.border = s["border"]
        ws3.row_dimensions[row].height = 18

    # --- Save ---
    date_str = datetime.date.today().strftime("%m-%d-%Y")
    filename = f"WorkOrder_{product.replace(' ', '_')}_{work_order}_{date_str}.xlsx"
    wb.save(filename)
    print(f"\nWork order saved as: {filename}")
    return filename


if __name__ == "__main__":
    product, speed, work_order, start_date, end_date, order_qty, shift_target = get_user_inputs()
    build_work_order(product, speed, work_order, start_date, end_date, order_qty, shift_target)
    print("\nDone! Open your work order to review.")