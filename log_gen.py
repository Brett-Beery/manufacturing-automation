import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_log_template(output_path="process_log_template.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Process Log"

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F8F")
    section_font = Font(name="Calibri", bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    normal_font = Font(name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Column Widths ---
    col_widths = [20, 25, 15, 15, 15, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Title Block ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "PROCESS LOG"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = "Manufacturing Automation Demo"
    ws["A2"].font = Font(name="Calibri", bold=True, size=12)
    ws["A2"].alignment = center

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # --- Info Block ---
    info_labels = ["Product:", "Line Speed:", "Date:", "Shift:", "Operator:", "Work Order #:"]
    for i, label in enumerate(info_labels):
        row = i + 3
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=2, value="").border = border
        ws.row_dimensions[row].height = 18

    # --- Raw Material Check Section ---
    rm_start = 10
    ws.merge_cells(f"A{rm_start}:F{rm_start}")
    ws[f"A{rm_start}"] = "RAW MATERIAL CHECKS"
    ws[f"A{rm_start}"].font = section_font
    ws[f"A{rm_start}"].fill = section_fill
    ws[f"A{rm_start}"].alignment = center

    rm_headers = ["Check Item", "Specification", "Min", "Max", "Actual", "Pass/Fail"]
    for col, header in enumerate(rm_headers, 1):
        cell = ws.cell(row=rm_start + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Blank rows for raw material checks
    for row in range(rm_start + 2, rm_start + 7):
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.row_dimensions[row].height = 18

    # --- Machine Parameter Check Section ---
    mp_start = rm_start + 8
    ws.merge_cells(f"A{mp_start}:F{mp_start}")
    ws[f"A{mp_start}"] = "MACHINE PARAMETER CHECKS"
    ws[f"A{mp_start}"].font = section_font
    ws[f"A{mp_start}"].fill = section_fill
    ws[f"A{mp_start}"].alignment = center

    mp_headers = ["Parameter", "Specification", "Min", "Max", "Actual", "Pass/Fail"]
    for col, header in enumerate(mp_headers, 1):
        cell = ws.cell(row=mp_start + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row in range(mp_start + 2, mp_start + 7):
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.row_dimensions[row].height = 18

    # --- Quality Summary Section ---
    qs_start = mp_start + 8
    ws.merge_cells(f"A{qs_start}:F{qs_start}")
    ws[f"A{qs_start}"] = "QUALITY SUMMARY"
    ws[f"A{qs_start}"].font = section_font
    ws[f"A{qs_start}"].fill = section_fill
    ws[f"A{qs_start}"].alignment = center

    ws.merge_cells(f"A{qs_start+1}:C{qs_start+1}")
    ws[f"A{qs_start+1}"] = "Total Checks:"
    ws.merge_cells(f"D{qs_start+1}:F{qs_start+1}")
    ws[f"D{qs_start+1}"] = f"=COUNTA(F{rm_start+2}:F{mp_start+6})"
    ws[f"D{qs_start+1}"].border = border

    ws.merge_cells(f"A{qs_start+2}:C{qs_start+2}")
    ws[f"A{qs_start+2}"] = "Total Pass:"
    ws.merge_cells(f"D{qs_start+2}:F{qs_start+2}")
    ws[f"D{qs_start+2}"] = f"=COUNTIF(F{rm_start+2}:F{mp_start+6},\"PASS\")"
    ws[f"D{qs_start+2}"].border = border

    ws.merge_cells(f"A{qs_start+3}:C{qs_start+3}")
    ws[f"A{qs_start+3}"] = "Total Fail:"
    ws.merge_cells(f"D{qs_start+3}:F{qs_start+3}")
    ws[f"D{qs_start+3}"] = f"=COUNTIF(F{rm_start+2}:F{mp_start+6},\"FAIL\")"
    ws[f"D{qs_start+3}"].border = border

    # --- Sign Off ---
    so_start = qs_start + 5
    ws.merge_cells(f"A{so_start}:F{so_start}")
    ws[f"A{so_start}"] = "SIGN OFF"
    ws[f"A{so_start}"].font = section_font
    ws[f"A{so_start}"].fill = section_fill
    ws[f"A{so_start}"].alignment = center

    signoff_labels = ["Operator Signature:", "Supervisor Signature:", "Date:"]
    for i, label in enumerate(signoff_labels):
        row = so_start + i + 1
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"C{row}:F{row}")
        ws[f"C{row}"].border = border
        ws.row_dimensions[row].height = 20

    wb.save(output_path)
    print(f"Template saved to {output_path}")

if __name__ == "__main__":
    create_log_template()